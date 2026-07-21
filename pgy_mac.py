"""蒲公英达人采集 macOS版 — ARM64原生
基于Windows版pgy_standalone.py (2026-07-14)，增加Chrome自动启动和Finder集成
打包: pyinstaller --onefile --console --hidden-import openpyxl --name 蒲公英采集 pgy_mac.py
"""
import websocket
import json
import urllib.request
import time
import csv
import os
import sys
import re
import subprocess
import threading
import openpyxl
import tkinter as tk
from tkinter import filedialog

# ====== 配置 ======
CDP_HOST = "127.0.0.1"
CDP_PORT = 9222
CHROME_PATHS = [
    "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
    "/Applications/Chromium.app/Contents/MacOS/Google Chrome",
    os.path.expanduser("~/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"),
]
OUT_DIR = os.path.dirname(os.path.abspath(sys.argv[0])) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))

def find_chrome():
    for p in CHROME_PATHS:
        if os.path.exists(p):
            return p
    return None

def start_chrome():
    chrome = find_chrome()
    if not chrome:
        print("未找到Chrome，请确认已安装Google Chrome")
        return False
    user_data = os.path.expanduser("~/chrome-auto-profile")
    try:
        subprocess.Popen([
            chrome,
            "--remote-debugging-port=9222",
            "--remote-allow-origins=*",
            f"--user-data-dir={user_data}",
        ], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        print("正在启动Chrome调试模式...")
        time.sleep(3)
        return True
    except Exception as e:
        print(f"启动Chrome失败: {e}")
        return False

def check_cdp():
    try:
        resp = urllib.request.urlopen(f"http://{CDP_HOST}:{CDP_PORT}/json", timeout=3)
        targets = json.loads(resp.read().decode('utf-8'))
        pages = [t for t in targets if t.get("type") == "page"]
        return len(pages) > 0
    except Exception:
        return False

def ask_excel_file():
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    root.lift()
    try:
        root.tk.call('tk', 'scaling', 2.0)  # Retina 适配
    except Exception:
        pass
    filepath = filedialog.askopenfilename(
        title="请选择达人列表Excel文件",
        filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
        initialdir=os.path.expanduser("~/Desktop")
    )
    root.destroy()
    return filepath

def read_ids_from_excel(path):
    """从Excel的Homepage列提取用户ID（表头行不固定，逐行搜索直至找到）"""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = None
    header_row = None
    homepage_col = None
    for sname in wb.sheetnames:
        s = wb[sname]
        if s.sheet_state == 'hidden':
            continue
        # 在所有行中搜索Homepage表头
        for row in range(1, s.max_row + 1):
            for col in range(1, s.max_column + 1):
                val = s.cell(row, col).value
                if val and 'homepage' in str(val).lower():
                    ws = s
                    header_row = row
                    homepage_col = col
                    break
            if ws:
                break
        if ws:
            break
    if not ws:
        raise ValueError("找不到包含Homepage列的sheet，请确认Excel格式")
    if not homepage_col:
        raise ValueError("找不到Homepage列，请确认Excel格式")

    user_ids = []
    for row in range(header_row + 1, ws.max_row + 1):
        url = ws.cell(row, homepage_col).value
        if url and 'user/profile/' in str(url):
            uid = str(url).split('user/profile/')[-1].split('?')[0].strip()
            if uid and uid not in user_ids:
                user_ids.append(uid)
    wb.close()
    return user_ids

sys.stdout.reconfigure(encoding='utf-8', errors='replace')


# ====== CDP客户端 ======
class CDP:
    def __init__(self):
        resp = urllib.request.urlopen(f"http://{CDP_HOST}:{CDP_PORT}/json")
        targets = json.loads(resp.read().decode('utf-8'))
        pages = [t for t in targets if t.get("type") == "page"]
        if not pages:
            raise RuntimeError("Chrome未启动或调试端口未开启")
        ws_url = pages[0]["webSocketDebuggerUrl"]

        self.ws = websocket.create_connection(ws_url, enable_multithread=True)
        self._id = 0
        self._results = {}
        self._events = []
        self._lock = threading.Lock()
        self._running = True
        self._recv = threading.Thread(target=self._recv_loop, daemon=True)
        self._recv.start()
        self.send("Page.enable")

    def _recv_loop(self):
        while self._running:
            try:
                msg = json.loads(self.ws.recv())
                mid = msg.get("id")
                if mid is not None:
                    with self._lock:
                        self._results[mid] = msg
                elif msg.get("method"):
                    self._events.append(msg)
            except Exception:
                if self._running:
                    time.sleep(0.1)

    def send(self, method, params=None, timeout=60):
        self._id += 1
        mid = self._id
        self.ws.send(json.dumps({"id": mid, "method": method, "params": params or {}}))
        start = time.time()
        while time.time() - start < timeout:
            with self._lock:
                if mid in self._results:
                    r = self._results.pop(mid)
                    if "error" in r:
                        raise RuntimeError(f"CDP error: {r['error']}")
                    return r.get("result", {})
            time.sleep(0.05)
        raise TimeoutError(f"CDP超时: {method}")

    def js(self, expression, await_promise=False):
        r = self.send("Runtime.evaluate", {
            "expression": expression,
            "awaitPromise": await_promise,
            "returnByValue": True
        })
        val = r.get("result", {}).get("value")
        if r.get("result", {}).get("type") == "object" and val is not None:
            return val
        if r.get("result", {}).get("subtype") == "error":
            raise RuntimeError(f"JS error: {r['result'].get('description','')}")
        return val

    def navigate_and_wait(self, url, wait=5):
        self.events_clear()
        self.send("Page.navigate", {"url": url})
        time.sleep(2)
        for _ in range(wait * 2):
            if any(e.get("method") == "Page.loadEventFired" for e in self._events):
                break
            time.sleep(0.5)
        time.sleep(1)

    def events_clear(self):
        self._events = []

    def current_url(self):
        return self.js("window.location.href")

    def close(self):
        self._running = False
        try:
            self.ws.close()
        except Exception:
            pass


# ====== 工具函数 ======
def fmt_pct(v):
    if v is None: return ""
    return f"{v*100:.1f}%" if isinstance(v, (int, float)) else f"{v}%"

def fmt_w(v): return f"{v/10000:.1f}w" if v else ""
def fmt_num(v): return f"{v:,}" if v else ""


# ====== Python API调用 ======
def api_get(cookie, url):
    req = urllib.request.Request(url)
    req.add_header("accept", "application/json, text/plain, */*")
    req.add_header("user-agent", "Mozilla/5.0")
    req.add_header("referer", "https://pgy.xiaohongshu.com/solar/pre-trade/blogger-detail/5a9c1b15e8ac2b1eb275ed95")
    req.add_header("cookie", cookie)
    with urllib.request.urlopen(req, timeout=15) as resp:
        return json.loads(resp.read().decode('utf-8'))

def api_post(cookie, url, payload):
    data = json.dumps(payload).encode('utf-8')
    req = urllib.request.Request(url, data=data)
    req.add_header("accept", "application/json, text/plain, */*")
    req.add_header("content-type", "application/json;charset=UTF-8")
    req.add_header("user-agent", "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36")
    req.add_header("referer", "https://pgy.xiaohongshu.com/solar/pre-trade/note/kol")
    req.add_header("origin", "https://pgy.xiaohongshu.com")
    req.add_header("cookie", cookie)
    with urllib.request.urlopen(req, timeout=15) as resp:
        return json.loads(resp.read().decode('utf-8'))


# ====== 主流程 ======
def main():
    print("=" * 50)
    print("蒲公英达人采集 macOS版")
    print("=" * 50)

    # Step 0: 确保Chrome调试模式运行
    print("\n检查Chrome调试端口...")
    if not check_cdp():
        print("Chrome调试端口未开启，尝试自动启动Chrome...")
        if start_chrome():
            for _ in range(10):
                if check_cdp():
                    print("Chrome调试端口已就绪")
                    break
                time.sleep(2)
            else:
                print("Chrome启动超时，请手动运行 启动Chrome_mac.command 后重试")
                try: input("按回车退出...")
                except EOFError: pass
                sys.exit(1)
        else:
            print("请先运行 启动Chrome_mac.command 启动Chrome调试模式")
            try: input("按回车退出...")
            except EOFError: pass
            sys.exit(1)
    else:
        print("Chrome调试端口已就绪")

    # Step 1: 选文件
    print("\n请选择达人列表Excel文件...")
    excel_file = ask_excel_file()
    if not excel_file:
        print("未选择文件，退出")
        try: input("按回车退出...")
        except EOFError: pass
        sys.exit(1)

    print(f"已选择: {os.path.basename(excel_file)}")
    try:
        user_ids = read_ids_from_excel(excel_file)
    except Exception as e:
        print(f"读取Excel失败: {e}")
        try: input("按回车退出...")
        except EOFError: pass
        sys.exit(1)

    if not user_ids:
        print("错误: 未从Excel中找到任何达人链接")
        try: input("按回车退出...")
        except EOFError: pass
        sys.exit(1)

    print(f"达人总数: {len(user_ids)}")

    # Step 2: 连接Chrome
    print("连接Chrome...")
    try:
        cdp = CDP()
    except Exception as e:
        print(f"连接失败: {e}")
        try: input("按回车退出...")
        except EOFError: pass
        sys.exit(1)

    try:
        # Step 3: 登录检测（用API实测，比URL检测更可靠）
        print("\n" + "=" * 50)
        print("登录检测")
        print("=" * 50)

        cdp.navigate_and_wait("https://pgy.xiaohongshu.com/solar/pre-trade")
        time.sleep(2)

        cookie = cdp.js("document.cookie", await_promise=False)
        test_ok = False
        try:
            test = api_get(cookie, f"https://pgy.xiaohongshu.com/api/solar/cooperator/user/blogger/{user_ids[0]}")
            if test.get("code") == 0:
                test_ok = True
                print("已登录")
        except Exception:
            pass

        if not test_ok:
            print("\n!! 未登录，请在Chrome窗口中手动登录蒲公英")
            print("!! 登录地址: https://pgy.xiaohongshu.com")
            print("!! 等待登录（每3秒检查一次，最多5分钟）...\n")
            logged_in = False
            for i in range(100):
                time.sleep(3)
                cookie = cdp.js("document.cookie", await_promise=False)
                try:
                    test = api_get(cookie, f"https://pgy.xiaohongshu.com/api/solar/cooperator/user/blogger/{user_ids[0]}")
                    if test.get("code") == 0:
                        logged_in = True
                        print("登录成功!")
                        break
                except Exception:
                    pass
                if i % 20 == 0 and i > 0:
                    print(f"仍在等待... ({i*3//60}分钟)")

            if not logged_in:
                print("登录超时，请重试")
                return

        cookie = cdp.js("document.cookie", await_promise=False)
        print(f"Cookie获取成功 (长度: {len(cookie)})")
        if not cookie or len(cookie) < 20:
            print("Cookie为空，请确认已登录蒲公英后再试")
            return

        # Step 4: 批量采集
        print("\n" + "=" * 50)
        print("批量采集 (API + notes_rate)")
        print("=" * 50)

        rows = []
        errors = []

        for i, uid in enumerate(user_ids):
            print(f"[{i+1}/{len(user_ids)}] {uid}...", end=" ", flush=True)

            # detail API
            try:
                d = api_get(cookie, f"https://pgy.xiaohongshu.com/api/solar/cooperator/user/blogger/{uid}")
                if d.get("code") != 0:
                    errors.append((uid, f"detail:{d.get('msg','')}"))
                    rows.append({"user_id": uid, "_error": f"detail failed: {d.get('msg','')}"})
                    print("FAIL detail")
                    continue
                d = d["data"]
                name = d.get("name", "")
                print(f"{name}", end="", flush=True)
            except Exception as e:
                errors.append((uid, str(e)))
                rows.append({"user_id": uid, "_error": str(e)})
                print(f"FAIL {e}")
                continue

            # fans_profile
            try:
                fans = api_get(cookie, f"https://pgy.xiaohongshu.com/api/solar/kol/data/{uid}/fans_profile")
                fd = fans.get("data", {}) if fans.get("code") == 0 else {}
            except Exception:
                fd = {}

            # v2 search
            v2 = {}
            if name:
                try:
                    sr = api_post(cookie, "https://pgy.xiaohongshu.com/api/solar/cooperator/blogger/v2",
                                  {"keyword": name, "pageNum": 1, "pageSize": 5})
                    if sr.get("code") == 0:
                        for k in sr["data"]["kols"]:
                            if k.get("userId") == uid:
                                v2 = k
                                break
                except Exception:
                    pass

            # notes_rate v3 (浏览器$http)
            nr = {}
            try:
                nr_raw = cdp.js(f"""
                    (async function() {{
                        try {{
                            const resp = await window.$http.get(
                                '/api/solar/kol/data_v3/notes_rate',
                                {{ params: {{ userId: '{uid}', business: 1, noteType: 3, dateType: 1, advertiseSwitch: 1 }} }}
                            );
                            return JSON.stringify(resp || {{}});
                        }} catch(e) {{
                            return JSON.stringify({{error: e.toString()}});
                        }}
                    }})()
                """, await_promise=True)
                nr = json.loads(nr_raw) if nr_raw else {}
                print("nr", end="", flush=True)
            except Exception as e:
                print(f"nrERR:{e}", end="")

            print()

            # 组装数据
            ct = [t.get("taxonomy1Tag","") for t in (d.get("contentTags") or [])]
            ft = d.get("featureTags") or []
            types_list = []
            if d.get("pictureState") == 1: types_list.append("图文")
            if d.get("videoState") == 1: types_list.append("视频")
            pic_p = d.get("picturePrice") or 0
            vid_p = d.get("videoPrice") or 0
            ages_sorted = sorted(fd.get("ages",[]), key=lambda x: x.get("percent",0) or 0, reverse=True)
            cities_list = (fd.get("cities") or [])[:5]

            # 观众画像
            gender_male = fd.get('gender',{}).get('male')
            gender_female = fd.get('gender',{}).get('female')
            gender_str = f"男:{fmt_pct(gender_male)} / 女:{fmt_pct(gender_female)}" if (gender_male is not None or gender_female is not None) else ""
            top2_cities = "、".join(f"{c['name']}({fmt_pct(c.get('percent'))})" for c in cities_list[:2])
            top2_ages = "、".join(f"{a['group']}({fmt_pct(a.get('percent'))})" for a in ages_sorted[:2])

            row = {
                "user_id": uid,
                "Platform（红书）": "小红书",
                "Hotel Name": "",
                "Influencer name": name,
                "Followers（w）": fmt_w(d.get("fansCount")),
                "Homepage": f"https://www.xiaohongshu.com/user/profile/{d.get('redId','')}",
                "达人所在地": d.get("location",""),
                "达人类型": "、".join(ct + (ft or [])),
                "合作形式（图文/视频）": "/".join(types_list) if types_list else "",
                "发布时间": "",
                "平台裸价（图文）": f"¥{pic_p:,.0f}" if pic_p else "",
                "平台裸价（视频）": f"¥{vid_p:,.0f}" if vid_p else "",
                "Platform service charge 星图/蒲公英": "",
                "Travel cost breakdown in detail": "",
                "Travel expense amount": "",
                "content authorization fee": "",
                "Offline event attendance fee": "",
                "specific social/EC platform distrubution fee": "",
                "Video operating expenses Feeds": "",
                "Video operating expenses SEM": "",
                "核心平台（红书）Total总价": "",
                "Packed total price": "",
                "Impression": "",
                "Views": "",
                "Engagement": "",
                "CPM": "",
                "CTR": "",
                "CPE": "",
                "曝光中位数": fmt_num(nr.get("impMedian")),
                "阅读中位数": fmt_num(nr.get("readMedian")),
                "互动中位数": fmt_num(nr.get("mengagementNum")),  # 页面显示用mengagementNum，不是interactionMedian（口径不同）
                "互动率": f"{nr.get('interactionRate')}%" if nr.get("interactionRate") is not None else "",
                "3s阅读率": f"{nr.get('picture3sViewRate')}%" if nr.get("picture3sViewRate") is not None else "",
                "完播率": f"{nr.get('videoFullViewRate')}%" if nr.get("videoFullViewRate") is not None else "",
                "活跃粉丝占比": f"{v2.get('fansActiveIn28dLv')}%" if v2.get('fansActiveIn28dLv') is not None else "",
                "观众画像-城市分布Top2": top2_cities,
                "观众画像-年龄分布Top2": top2_ages,
                "观众画像-男女比例": gender_str,
                "权益（授权期限，授权渠道，集团授权，酒店分发授权，尽可能补充全部分）": "",
                "内容可供免费二次剪辑和混剪，授权文件需提供无bgm无字幕纯净版素材和最终发布版素材": "",
                "可接受发布添加搜索词组件": "",
                "可接受发布添加评论区组件": "",
                "推荐理由": "",
            }
            rows.append(row)

            if (i + 1) % 10 == 0:
                cp_path = os.path.join(OUT_DIR, "_checkpoint.json")
                with open(cp_path, "w", encoding="utf-8") as f:
                    json.dump(rows, f, ensure_ascii=False, indent=2)
                print(f"  [checkpoint: {i+1}/{len(user_ids)}]")

            time.sleep(0.3)

        # 输出
        print("\n" + "=" * 50)
        print("输出结果")
        print("=" * 50)

        json_path = os.path.join(OUT_DIR, "达人数据_v3_结果.json")
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(rows, f, ensure_ascii=False, indent=2)

        csv_path = os.path.join(OUT_DIR, "达人数据_v3.csv")
        fieldnames = [
            "Platform（红书）","Hotel Name","Influencer name","Followers（w）","Homepage",
            "达人所在地","达人类型","合作形式（图文/视频）","发布时间",
            "平台裸价（图文）","平台裸价（视频）",
            "Platform service charge 星图/蒲公英",
            "Travel cost breakdown in detail","Travel expense amount",
            "content authorization fee","Offline event attendance fee",
            "specific social/EC platform distrubution fee",
            "Video operating expenses Feeds","Video operating expenses SEM",
            "核心平台（红书）Total总价","Packed total price",
            "Impression","Views","Engagement","CPM","CTR","CPE",
            "曝光中位数","阅读中位数","互动中位数","互动率",
            "3s阅读率","完播率","活跃粉丝占比",
            "观众画像-城市分布Top2","观众画像-年龄分布Top2","观众画像-男女比例",
            "权益（授权期限，授权渠道，集团授权，酒店分发授权，尽可能补充全部分）",
            "内容可供免费二次剪辑和混剪，授权文件需提供无bgm无字幕纯净版素材和最终发布版素材",
            "可接受发布添加搜索词组件","可接受发布添加评论区组件",
            "推荐理由",
        ]
        with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction='ignore')
            writer.writeheader()
            writer.writerows(rows)

        ok = sum(1 for r in rows if not r.get("_error"))
        print(f"完成: {ok}/{len(rows)} 成功")
        if errors:
            print(f"错误 ({len(errors)}):")
            for uid, err in errors[:5]:
                print(f"  {uid}: {err}")
        print(f"\nCSV: {csv_path}")

        # 清理临时文件
        for tmp in ["_checkpoint.json"]:
            p = os.path.join(OUT_DIR, tmp)
            if os.path.exists(p):
                os.remove(p)

        # macOS: 在Finder中定位输出文件
        try:
            subprocess.run(["open", "-R", csv_path])
        except Exception:
            pass

    finally:
        cdp.close()

    print("\n按回车退出...")
    try:
        input()
    except EOFError:
        pass


if __name__ == "__main__":
    main()
